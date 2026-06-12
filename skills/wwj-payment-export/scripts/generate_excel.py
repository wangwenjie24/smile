#!/usr/bin/env python3
"""
根据 JSON 生成「预计支付款项明细表」Excel。
用法: python generate_excel.py <input.json> <output.xlsx>
"""
import json, sys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def gen(data, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "预计支付款项明细表"

    # 样式
    F = lambda sz=11, b=False: Font(name="宋体", size=sz, bold=b)
    AC = Alignment(horizontal="center", vertical="center")
    AL = Alignment(horizontal="left", vertical="center")
    AR = Alignment(horizontal="right", vertical="center")
    BD = Border(*(Side(style="thin"),) * 4)
    HDR_FILL = PatternFill("solid", fgColor="D6DCE4")  # 浅蓝偏灰

    for col, w in {1: 10, 2: 55, 3: 18, 4: 18, 5: 30, 6: 15}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Row 1: 标题
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "预计支付款项明细表"
    c.font = F(16, True)
    c.alignment = AC

    # Row 2: 支付期间
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = f"支付期间：{date.today().strftime('%Y年%m月%d日')}"
    c.font = F()
    c.alignment = AL

    # Row 3: 表头（带背景色）
    headers = ["部门", "项目", "计划付款金额（元）", "实际付款金额（元）", "备注", "内容"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = F(b=True)
        c.alignment = AC
        c.border = BD
        c.fill = HDR_FILL
    ws.row_dimensions[3].height = 28.8

    # 数据行
    # 对齐规则：A部门=center, B项目=left, C金额=right, D实际付款=right, E备注=left, F内容=center
    col_align = {1: AC, 2: AL, 3: AR, 4: AR, 5: AL, 6: AC}
    items = data.get("data", [])
    for i, it in enumerate(items):
        r = 4 + i
        vals = ["", it.get("project", ""), it.get("amount", ""), "", it.get("remark", ""), ""]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=j)
            if j == 3:  # 金额列
                try:
                    c.value = float(v)
                    c.number_format = "#,##0.00"
                except (ValueError, TypeError):
                    c.value = v
            else:
                c.value = v
            c.font = F()
            c.alignment = col_align[j]
            c.border = BD

    # 合计行
    tr = 4 + len(items)
    ws.merge_cells(f"A{tr}:B{tr}")
    c = ws.cell(row=tr, column=1, value="合计")
    c.font = F(b=True)
    c.alignment = AR
    c.border = BD
    ws.cell(row=tr, column=2).border = BD

    c = ws.cell(row=tr, column=3)
    c.value = f"=SUM(C4:C{tr - 1})" if items else 0
    c.font = F(b=True)
    c.alignment = AR
    c.border = BD
    c.number_format = "#,##0.00"

    for col in [4, 5, 6]:
        c = ws.cell(row=tr, column=col)
        c.border = BD
        c.font = F(b=True)

    wb.save(path)
    print(f"已生成: {path}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python generate_excel.py <input.json> <output.xlsx>")
        sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as f:
        gen(json.load(f), sys.argv[2])
